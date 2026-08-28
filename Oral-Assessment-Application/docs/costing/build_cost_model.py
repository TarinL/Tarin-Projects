from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

ARIAL = "Arial"
BLUE = Font(name=ARIAL, color="0000FF")          # hardcoded inputs
BLACK = Font(name=ARIAL, color="000000")         # formulas
GREEN = Font(name=ARIAL, color="008000")         # cross-sheet links
BOLD = Font(name=ARIAL, bold=True)
TITLE = Font(name=ARIAL, bold=True, size=14)
HDR = Font(name=ARIAL, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2563EB")
SEC_FILL = PatternFill("solid", fgColor="DBEAFE")
YEL = PatternFill("solid", fgColor="FFFF00")
NZD = '"$"#,##0.00'
NZD3 = '"$"#,##0.000'
USD = '"$"#,##0.0000'
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ============================ INPUTS ============================
ws = wb.active
ws.title = "Inputs"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 38
for c in "BCDE":
    ws.column_dimensions[c].width = 16

def put(cell, val, font=BLACK, fmt=None, fill=None, align=None, comment=None):
    ws[cell] = val
    ws[cell].font = font
    if fmt: ws[cell].number_format = fmt
    if fill: ws[cell].fill = fill
    if align: ws[cell].alignment = Alignment(horizontal=align)
    if comment: ws[cell].comment = Comment(comment, "cost-model")

def section(row, text):
    ws[f"A{row}"] = text
    ws[f"A{row}"].font = BOLD
    for c in "ABCDE":
        ws[f"{c}{row}"].fill = SEC_FILL

put("A1", "Interview Cost Model — Inputs", TITLE)
put("A2", "All blue cells are editable. Black = formulas. Currency: NZD unless marked USD.", Font(name=ARIAL, italic=True, size=9))

section(3, "SCENARIO CONTROLS (edit these)")
put("A4", "Interview length (minutes)"); put("B4", 10, BLUE, fill=YEL)
put("A5", "Depth tier (1=Shallow, 2=Standard, 3=Deep)"); put("B5", 1, BLUE, fill=YEL,
    comment="1 = current config (max_tokens 120, gpt-4o marker)\n2 = max_tokens 300, gpt-4o\n3 = max_tokens 800, reasoning marker")
put("A6", "Students per class"); put("B6", 30, BLUE, fill=YEL)
put("A7", "Assessments per semester"); put("B7", 2, BLUE, fill=YEL)

section(9, "FX")
put("A10", "USD -> NZD rate"); put("B10", 1.71, BLUE, fmt="0.000", fill=YEL,
    comment="Spot ~1.71 (11 Jun 2026). Update as needed.")

section(12, "UNIT PRICES (USD, list)")
put("A13", "gpt-4o input ($/1M tok)"); put("B13", 2.50, BLUE, fmt=USD)
put("A14", "gpt-4o output ($/1M tok)"); put("B14", 10.00, BLUE, fmt=USD)
put("A15", "gpt-4o-mini input ($/1M tok)"); put("B15", 0.15, BLUE, fmt=USD)
put("A16", "gpt-4o-mini output ($/1M tok)"); put("B16", 0.60, BLUE, fmt=USD)
put("A17", "reasoning input ($/1M tok)"); put("B17", 2.00, BLUE, fmt=USD)
put("A18", "reasoning output ($/1M tok)"); put("B18", 8.00, BLUE, fmt=USD)
put("A19", "ElevenLabs ($/character)"); put("B19", 0.00011, BLUE, fmt='0.000000', fill=YEL,
    comment="HIGHEST-LEVERAGE ASSUMPTION. Turbo v2.5 = 0.5 credit/char; Creator $22/100k credits -> ~$0.00011/char. Pro/Scale cheaper.")
put("A20", "recall.ai ($/bot-minute)"); put("B20", 0.00833, BLUE, fmt='0.00000',
    comment="$0.50 / recording-hour pay-as-you-go (2026 price), prorated per second.")
put("A21", "Fargate ($/vCPU-hour, Sydney)"); put("B21", 0.04856, BLUE, fmt='0.00000')
put("A22", "Fargate ($/GB-hour, Sydney)"); put("B22", 0.00532, BLUE, fmt='0.00000')

section(24, "TASK / USAGE ASSUMPTIONS")
put("A25", "Fargate vCPUs"); put("B25", 8, BLUE, comment="task_definition.json cpu=8192 (rev :7)")
put("A26", "Fargate memory (GB)"); put("B26", 16, BLUE, comment="task_definition.json memory=16384 (rev :7)")
put("A27", "Fargate cold-start overhead (min)"); put("B27", 3, BLUE,
    comment="Image pull + whisper load + post-interview marking, beyond the live interview.")
put("A28", "Interviewer LLM calls (per baseline interview)"); put("B28", 13, BLUE)
put("A29", "Interviewer avg input tok / call"); put("B29", 1500, BLUE)
put("A30", "Marking input tokens"); put("B30", 6000, BLUE,
    comment="Measured 2026-06-12: live gpt-4o marking call on a full 10-min transcript = 5,998 in / 442 out.")
put("A31", "TTS characters (per baseline interview)"); put("B31", 2850, BLUE)
put("A32", "Baseline minutes (for per-minute scaling)"); put("B32", 10, BLUE,
    comment="Time-scaled usage (TTS chars, LLM calls) is divided by this to get a per-minute rate.")

section(34, "DEPTH TIER TABLE")
for col, txt in zip("ABCDE", ["Tier", "Interviewer max_tokens", "Marking output tok", "Marker in $/1M", "Marker out $/1M"]):
    put(f"{col}35", txt, HDR, fill=HDR_FILL, align="center")
# tier rows 36-38
tiers = [(1,120,500,"=B13","=B14"), (2,300,2500,"=B13","=B14"), (3,800,6000,"=B17","=B18")]
for i,(t,mt,mo,din,dout) in enumerate(tiers):
    r = 36+i
    put(f"A{r}", t, BLUE, align="center")
    put(f"B{r}", mt, BLUE, align="center")
    put(f"C{r}", mo, BLUE, align="center")
    put(f"D{r}", din, BLACK, fmt=USD, align="center")
    put(f"E{r}", dout, BLACK, fmt=USD, align="center")

section(40, "SELECTED DEPTH (from tier control B5)")
put("A41", "Interviewer max_tokens"); put("B41", "=INDEX(B36:B38,$B$5)", BLACK, align="center")
put("A42", "Marking output tokens"); put("B42", "=INDEX(C36:C38,$B$5)", BLACK, align="center")
put("A43", "Marker input $/1M"); put("B43", "=INDEX(D36:D38,$B$5)", BLACK, fmt=USD, align="center")
put("A44", "Marker output $/1M"); put("B44", "=INDEX(E36:E38,$B$5)", BLACK, fmt=USD, align="center")

# ============================ PER-INTERVIEW ============================
pi = wb.create_sheet("Per-Interview")
pi.sheet_view.showGridLines = False
pi.column_dimensions["A"].width = 40
for c in "BCDE":
    pi.column_dimensions[c].width = 15

def p(cell, val, font=BLACK, fmt=None, fill=None, align=None, bold=False):
    pi[cell] = val
    pi[cell].font = Font(name=ARIAL, bold=bold, color=font.color.rgb if font.color else "000000")
    if fmt: pi[cell].number_format = fmt
    if fill: pi[cell].fill = fill
    if align: pi[cell].alignment = Alignment(horizontal=align)

pi["A1"] = "Per-Interview Cost"; pi["A1"].font = TITLE
pi["A3"] = "Selected scenario"; pi["A3"].font = BOLD
for c in "ABCDE": pi[f"{c}3"].fill = SEC_FILL
p("A4", "Interview minutes"); p("B4", "=Inputs!B4", GREEN, align="center")
p("A5", "Depth tier"); p("B5", "=Inputs!B5", GREEN, align="center")
p("A6", "USD -> NZD"); p("B6", "=Inputs!B10", GREEN, fmt="0.000", align="center")

# header
for col, txt in zip("ABCDE", ["Component", "Scales with", "USD", "NZD", "Share"]):
    p(f"{col}8", txt, HDR, fill=HDR_FILL, align="center" if col!="A" else "left")

# Fargate $/min expression reused
FARG_MIN = "(Inputs!B25*Inputs!B21+Inputs!B26*Inputs!B22)/60"

rows = [
    ("ElevenLabs TTS", "time", f"=(Inputs!B31/Inputs!B32*$B$4)*Inputs!B19"),
    ("recall.ai bot", "time", f"=Inputs!B20*$B$4"),
    ("Fargate (8 vCPU / 16 GB)", "time", f"={FARG_MIN}*($B$4+Inputs!B27)"),
    ("OpenAI marking (gpt-4o)", "depth", f"=(Inputs!B30*Inputs!B43+Inputs!B42*Inputs!B44)/1000000"),
    ("OpenAI interviewer (gpt-4o-mini)", "time + depth", f"=(Inputs!B28/Inputs!B32*$B$4)*(Inputs!B29*Inputs!B15+Inputs!B41*Inputs!B16)/1000000"),
    ("faster-whisper STT (local)", "-", "=0"),
]
start = 9
for i,(name, scal, usd) in enumerate(rows):
    r = start+i
    p(f"A{r}", name)
    p(f"B{r}", scal, align="center")
    p(f"C{r}", usd, fmt=USD)
    p(f"D{r}", f"=C{r}*$B$6", fmt=NZD)
    p(f"E{r}", f"=D{r}/$D${start+len(rows)}", fmt="0.0%", align="center")
    for c in "ABCDE": pi[f"{c}{r}"].border = BORDER

tot = start+len(rows)
p(f"A{tot}", "Total per interview", bold=True)
p(f"C{tot}", f"=SUM(C{start}:C{tot-1})", fmt=USD, bold=True)
p(f"D{tot}", f"=SUM(D{start}:D{tot-1})", fmt=NZD, bold=True)
p(f"E{tot}", f"=SUM(E{start}:E{tot-1})", fmt="0.0%", bold=True, align="center")
for c in "ABCDE":
    pi[f"{c}{tot}"].border = BORDER
    pi[f"{c}{tot}"].fill = PatternFill("solid", fgColor="DCFCE7")

# headline derived metrics
h = tot+2
pi[f"A{h}"] = "Headline metrics"; pi[f"A{h}"].font = BOLD
for c in "ABCDE": pi[f"{c}{h}"].fill = SEC_FILL
p(f"A{h+1}", "Per-interview cost (NZD)"); p(f"B{h+1}", f"=D{tot}", BLACK, fmt=NZD, bold=True)
# base rate = time-variable NZD / minutes (TTS+recall+Fargate-running+interviewer)
p(f"A{h+2}", "Base rate (NZD / interview-minute)")
p(f"B{h+2}", f"=(D{start}+D{start+1}+{FARG_MIN}*$B$4*$B$6+D{start+4})/$B$4", BLACK, fmt=NZD3)
p(f"A{h+3}", "Marking alone (NZD)"); p(f"B{h+3}", f"=D{start+3}", BLACK, fmt=NZD)
p(f"A{h+5}", "Note: base rate excludes marking & Fargate cold-start (the non-time-scaled pieces).",
  Font(name=ARIAL, italic=True, size=9))

# ============================ SCALING ============================
sc = wb.create_sheet("Scaling")
sc.sheet_view.showGridLines = False
sc.column_dimensions["A"].width = 34
for c in "BCDE":
    sc.column_dimensions[c].width = 16

def s(cell, val, font=BLACK, fmt=None, fill=None, align=None, bold=False):
    sc[cell] = val
    sc[cell].font = Font(name=ARIAL, bold=bold, color=font.color.rgb if font.color else "000000")
    if fmt: sc[cell].number_format = fmt
    if fill: sc[cell].fill = fill
    if align: sc[cell].alignment = Alignment(horizontal=align)

sc["A1"] = "Scaling — class & semester totals"; sc["A1"].font = TITLE
sc["A3"] = "Current scenario (uses Inputs tier & minutes)"; sc["A3"].font = BOLD
for c in "ABCDE": sc[f"{c}3"].fill = SEC_FILL
s("A4", "Per interview (NZD)"); s("B4", "='Per-Interview'!B%d" % (h+1), GREEN, fmt=NZD, bold=True)
s("A5", "Students per class"); s("B5", "=Inputs!B6", GREEN, align="center")
s("A6", "Assessments per semester"); s("B6", "=Inputs!B7", GREEN, align="center")
s("A7", "Per class, 1 assessment"); s("B7", "=B4*B5", BLACK, fmt=NZD)
s("A8", "Per class, full semester"); s("B8", "=B4*B5*B6", BLACK, fmt=NZD)

# tier parameters (NZD) for the matrix
sc["A11"] = "Tier parameters (NZD, independent of scenario)"; sc["A11"].font = BOLD
for c in "ABCDE": sc[f"{c}11"].fill = SEC_FILL
for col, txt in zip("BCD", ["Tier 1", "Tier 2", "Tier 3"]):
    s(f"{col}12", txt, HDR, fill=HDR_FILL, align="center")
FARG_MIN_S = "(Inputs!B25*Inputs!B21+Inputs!B26*Inputs!B22)/60"
# per-minute NZD per tier (TTS + recall + Fargate-running + interviewer); interviewer uses tier maxtok in B36..B38
s("A13", "Variable rate (NZD / min)")
for col, trow in zip("BCD", [36,37,38]):
    expr = (f"=((Inputs!B31/Inputs!B32)*Inputs!B19"
            f"+Inputs!B20"
            f"+{FARG_MIN_S}"
            f"+(Inputs!B28/Inputs!B32)*(Inputs!B29*Inputs!B15+Inputs!B{trow}*Inputs!B16)/1000000"
            f")*Inputs!B10")
    s(f"{col}13", expr, BLACK, fmt=NZD3)
# fixed NZD per tier (Fargate cold-start + marking, marking uses tier rows C/D/E 36..38)
s("A14", "Fixed per interview (NZD)")
for col, trow in zip("BCD", [36,37,38]):
    expr = (f"=({FARG_MIN_S}*Inputs!B27"
            f"+(Inputs!B30*Inputs!D{trow}+Inputs!C{trow}*Inputs!E{trow})/1000000"
            f")*Inputs!B10")
    s(f"{col}14", expr, BLACK, fmt=NZD3)

# matrix: per-interview NZD by minutes x depth
sc["A16"] = "Per-interview cost (NZD) by length x depth"; sc["A16"].font = BOLD
for c in "ABCDE": sc[f"{c}16"].fill = SEC_FILL
for col, txt in zip("ABCD", ["Minutes", "Shallow", "Standard", "Deep"]):
    s(f"{col}17", txt, HDR, fill=HDR_FILL, align="center")
for i, mins in enumerate([5,10,15,20]):
    r = 18+i
    s(f"A{r}", mins, BLUE, align="center")
    for col in "BCD":
        s(f"{col}{r}", f"=$A{r}*{col}$13+{col}$14", BLACK, fmt=NZD)
        sc[f"{col}{r}"].border = BORDER
    sc[f"A{r}"].border = BORDER

# course totals by depth (full semester, current students/assessments, 10-min)
sc["A24"] = "Semester cost (NZD) — class of N, all assessments, by depth (10-min interview)"; sc["A24"].font = BOLD
for c in "ABCDE": sc[f"{c}24"].fill = SEC_FILL
for col, txt in zip("BCD", ["Shallow", "Standard", "Deep"]):
    s(f"{col}25", txt, HDR, fill=HDR_FILL, align="center")
s("A26", "Per class, full semester")
for col in "BCD":
    # per-interview at 10 min for that tier * students * assessments
    s(f"{col}26", f"=(10*{col}13+{col}14)*Inputs!B6*Inputs!B7", BLACK, fmt=NZD)

wb.save("/Users/marcusfindlow/Projects/399/docs/costing/interview_cost_model.xlsx")
print("saved")
